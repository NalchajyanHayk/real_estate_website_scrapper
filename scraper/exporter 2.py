"""Excel workbook export and presentation."""

from datetime import datetime, timezone
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .config import ScraperConfig
from .models import CompanyRecord, ScrapeSummary

SHEET_NAME = "Real Estate Developers"
SUMMARY_SHEET = "Scrape Summary"


def export_to_excel(records: list[CompanyRecord], output_path: Path,
                    config: ScraperConfig | None = None,
                    summary: ScrapeSummary | None = None) -> Path:
    """Write records and run metadata to a formatted XLSX workbook."""
    config = config or ScraperConfig(output_path=output_path)
    summary = summary or ScrapeSummary(company_links_found=len(records), profiles_successfully_parsed=len(records))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    columns = list(CompanyRecord().to_flat_dict())
    frame = pd.DataFrame([record.to_flat_dict() for record in records], columns=columns)
    note = ""
    if summary.company_links_found < config.max_rows:
        note = f"Source offered {summary.company_links_found} unique profile links, fewer than configured limit {config.max_rows}; no rows were fabricated."
    summary_rows = [
        ("Source Category URL", config.start_url),
        ("Scrape Timestamp (UTC)", datetime.now(timezone.utc).replace(microsecond=0).isoformat()),
        ("Configured Maximum Rows", config.max_rows),
        ("Available Profile Links", summary.company_links_found),
        ("Successful Rows", len(records)),
        ("Failed Rows", summary.profiles_failed),
        ("Duplicate Count", summary.duplicates_removed),
        ("Note", note),
    ]
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        frame.to_excel(writer, sheet_name=SHEET_NAME, index=False)
        pd.DataFrame(summary_rows, columns=["Metric", "Value"]).to_excel(writer, sheet_name=SUMMARY_SHEET, index=False)
    workbook = load_workbook(output_path)
    sheet = workbook[SHEET_NAME]
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    widths = [30, 30, 45, 24, 22, 28, 28, 35, 45, 45, 35, 20, 45, 26]
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        row[5].number_format = "@"
        if row[12].value:
            row[12].hyperlink = str(row[12].value)
            row[12].style = "Hyperlink"
        if row[7].value and ";" not in str(row[7].value):
            row[7].hyperlink = str(row[7].value)
            row[7].style = "Hyperlink"
    summary_sheet = workbook[SUMMARY_SHEET]
    summary_sheet.freeze_panes = "A2"
    summary_sheet.auto_filter.ref = summary_sheet.dimensions
    summary_sheet.column_dimensions["A"].width = 30
    summary_sheet.column_dimensions["B"].width = 100
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
    for row in summary_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    workbook.save(output_path)
    return output_path
